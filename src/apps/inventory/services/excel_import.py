# ==========================================================
# SERVICIO DE IMPORTACIÓN MASIVA DE ACTIVOS
# SPRINT 16
# ==========================================================

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.utils.dateparse import parse_date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from apps.accounts.models import User
from apps.inventory.models import Asset, AcquisitionBatch


# ==========================================================
# COLUMNAS DE LA PLANTILLA
# ==========================================================

TEMPLATE_COLUMNS = [
    "codigo_interno",
    "codigo_patrimonial",
    "tipo_activo",
    "marca",
    "modelo",
    "numero_serie",
    "codigo_lote",
    "hostname",
    "usuario_email",
    "departamento",
    "ubicacion",
    "estado_operativo",
    "estado_conexion",
    "sistema_operativo",
    "direccion_ip",
    "direccion_mac",
    "fecha_compra",
    "vencimiento_garantia",
    "proveedor",
    "observaciones",
]


REQUIRED_COLUMNS = [
    "codigo_interno",
    "tipo_activo",
    "marca",
    "modelo",
    "codigo_patrimonial",
    "numero_serie",
    "codigo_lote",
]


# ==========================================================
# RESULTADOS
# ==========================================================

@dataclass
class ImportRowResult:
    row_number: int
    data: dict[str, Any]
    is_valid: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    total_rows: int = 0
    valid_rows: int = 0
    invalid_rows: int = 0
    imported_rows: int = 0
    skipped_rows: int = 0
    rows: list[ImportRowResult] = field(default_factory=list)

    @property
    def has_errors(self):
        return self.invalid_rows > 0


# ==========================================================
# FUNCIONES AUXILIARES
# ==========================================================

def normalize_header(value):
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )


def normalize_text(value):
    if value is None:
        return ""

    return str(value).strip()


def normalize_upper(value):
    return normalize_text(value).upper()


def normalize_mac_address(value):
    mac_address = normalize_upper(value).replace("-", ":")

    if not mac_address:
        return ""

    parts = mac_address.split(":")

    if len(parts) != 6 or any(len(part) != 2 for part in parts):
        raise ValueError(
            "La dirección MAC no tiene un formato válido."
        )

    try:
        int("".join(parts), 16)
    except ValueError as exc:
        raise ValueError(
            "La dirección MAC contiene caracteres inválidos."
        ) from exc

    return ":".join(parts)


def normalize_date(value):
    if value in (None, ""):
        return None

    if hasattr(value, "date"):
        return value.date()

    if hasattr(value, "year") and hasattr(value, "month"):
        return value

    text = normalize_text(value)

    parsed = parse_date(text)

    if parsed:
        return parsed

    for separator in ("/", "-"):
        parts = text.split(separator)

        if len(parts) == 3:
            try:
                day, month, year = map(int, parts)

                from datetime import date

                return date(year, month, day)

            except (TypeError, ValueError):
                pass

    raise ValueError(
        "Fecha inválida. Use AAAA-MM-DD o DD/MM/AAAA."
    )


def normalize_decimal(value):
    if value in (None, ""):
        return None

    text = normalize_text(value).replace(".", "").replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(
            "El valor numérico no es válido."
        ) from exc


def get_choice_map(model_field_name):
    field_object = Asset._meta.get_field(model_field_name)

    choice_map = {}

    for stored_value, label in field_object.choices:
        choice_map[normalize_upper(stored_value)] = stored_value
        choice_map[normalize_upper(label)] = stored_value

    return choice_map


def resolve_choice(value, field_name, field_label):
    normalized_value = normalize_upper(value)

    if not normalized_value:
        return None

    choice_map = get_choice_map(field_name)

    resolved = choice_map.get(normalized_value)

    if resolved is None:
        valid_labels = sorted(
            {
                str(label)
                for _, label in Asset._meta.get_field(field_name).choices
            }
        )

        raise ValueError(
            f"{field_label} inválido. Valores permitidos: "
            f"{', '.join(valid_labels)}."
        )

    return resolved


def find_user(email):
    email = normalize_text(email)

    if not email:
        return None

    return User.objects.filter(
        email__iexact=email
    ).first()


def is_empty_row(values):
    return all(
        value in (None, "")
        or normalize_text(value) == ""
        for value in values
    )


# ==========================================================
# LECTURA DEL ARCHIVO
# ==========================================================

def read_excel_file(uploaded_file):
    try:
        workbook = load_workbook(
            filename=uploaded_file,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValidationError(
            "No se pudo abrir el archivo Excel. "
            "Verifique que sea un archivo .xlsx válido."
        ) from exc

    worksheet = workbook.active

    header_values = [
        normalize_header(cell.value)
        for cell in worksheet[1]
    ]

    if not any(header_values):
        raise ValidationError(
            "El archivo no contiene encabezados."
        )

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in header_values
    ]

    if missing_columns:
        raise ValidationError(
            "Faltan columnas obligatorias: "
            + ", ".join(missing_columns)
        )

    return workbook, worksheet, header_values


# ==========================================================
# VALIDACIÓN DE FILAS
# ==========================================================

def validate_asset_row(
    row_number,
    raw_data,
    codes_seen=None,
    serials_seen=None,
):
    codes_seen = codes_seen if codes_seen is not None else set()
    serials_seen = serials_seen if serials_seen is not None else set()

    errors = []
    warnings = []

    internal_code = normalize_upper(
        raw_data.get("codigo_interno")
    )

    patrimonial_code = normalize_upper(
        raw_data.get("codigo_patrimonial")
    )

    serial_number = normalize_upper(
        raw_data.get("numero_serie")
    )
    batch_code = normalize_upper(raw_data.get("codigo_lote"))

    if not internal_code:
        errors.append(
            "El código interno es obligatorio."
        )

    elif internal_code in codes_seen:
        errors.append(
            "El código interno está duplicado dentro del archivo."
        )

    elif Asset.objects.filter(
        internal_code__iexact=internal_code
    ).exists():
        errors.append(
            "El código interno ya existe en el inventario."
        )

    if serial_number:
        if serial_number in serials_seen:
            errors.append(
                "El número de serie está duplicado dentro del archivo."
            )

        elif Asset.objects.filter(
            serial_number__iexact=serial_number
        ).exists():
            errors.append(
                "El número de serie ya existe en el inventario."
            )

    if patrimonial_code and Asset.objects.filter(
        patrimonial_code__iexact=patrimonial_code
    ).exists():
        errors.append(
            "El código patrimonial ya existe en el inventario."
        )

    for field_label, field_value in {
        "La marca": raw_data.get("marca"),
        "El modelo": raw_data.get("modelo"),
        "El código patrimonial": patrimonial_code,
        "El número de serie": serial_number,
        "El código de lote": batch_code,
    }.items():
        if not normalize_text(field_value):
            errors.append(f"{field_label} es obligatorio.")

    acquisition_batch = None
    if batch_code:
        acquisition_batch = AcquisitionBatch.objects.filter(code__iexact=batch_code).first()
        if not acquisition_batch:
            errors.append(f"No existe el lote {batch_code}.")
        elif acquisition_batch.status not in {
            AcquisitionBatch.Status.VALIDATED,
            AcquisitionBatch.Status.CLOSED,
        }:
            errors.append(f"El lote {batch_code} todavía no está validado.")
        elif not acquisition_batch.audit_documents.filter(verified=True).exists():
            errors.append(f"El lote {batch_code} no tiene documentación verificada.")

    try:
        asset_type = resolve_choice(
            raw_data.get("tipo_activo"),
            "asset_type",
            "Tipo de activo",
        )
    except ValueError as exc:
        asset_type = None
        errors.append(str(exc))

    try:
        operational_status = resolve_choice(
            raw_data.get("estado_operativo"),
            "operational_status",
            "Estado operativo",
        )
    except ValueError as exc:
        operational_status = None
        errors.append(str(exc))

    try:
        connection_status = resolve_choice(
            raw_data.get("estado_conexion"),
            "connection_status",
            "Estado de conexión",
        )
    except ValueError as exc:
        connection_status = None
        errors.append(str(exc))

    user_email = normalize_text(
        raw_data.get("usuario_email")
    )

    assigned_user = find_user(user_email)

    if user_email and assigned_user is None:
        errors.append(
            f"No existe un usuario activo con el correo {user_email}."
        )

    try:
        mac_address = normalize_mac_address(
            raw_data.get("direccion_mac")
        )
    except ValueError as exc:
        mac_address = ""
        errors.append(str(exc))

    try:
        purchase_date = normalize_date(
            raw_data.get("fecha_compra")
        )
    except ValueError as exc:
        purchase_date = None
        errors.append(
            f"Fecha de compra: {exc}"
        )

    try:
        warranty_expiration = normalize_date(
            raw_data.get("vencimiento_garantia")
        )
    except ValueError as exc:
        warranty_expiration = None
        errors.append(
            f"Vencimiento de garantía: {exc}"
        )

    if (
        purchase_date
        and warranty_expiration
        and warranty_expiration < purchase_date
    ):
        errors.append(
            "La garantía no puede vencer antes de la fecha de compra."
        )

    cleaned_data = {
        "internal_code": internal_code,
        "patrimonial_code": patrimonial_code,
        "asset_type": asset_type,
        "brand": normalize_text(raw_data.get("marca")),
        "model": normalize_text(raw_data.get("modelo")),
        "serial_number": serial_number,
        "hostname": normalize_upper(raw_data.get("hostname")),
        "acquisition_batch": acquisition_batch,
        "assigned_user": assigned_user,
        "department": normalize_text(
            raw_data.get("departamento")
        ),
        "location": normalize_text(
            raw_data.get("ubicacion")
        ),
        "operational_status": operational_status,
        "connection_status": connection_status,
        "operating_system": normalize_text(
            raw_data.get("sistema_operativo")
        ),
        "current_ip": normalize_text(
            raw_data.get("direccion_ip")
        ),
        "mac_address": mac_address,
        "purchase_date": purchase_date,
        "warranty_expiration": warranty_expiration,
        "supplier": normalize_text(
            raw_data.get("proveedor")
        ),
        "notes": normalize_text(
            raw_data.get("observaciones")
        ),
    }

    # Los campos con valores predeterminados del modelo se omiten
    # cuando el Excel no contiene un valor.
    if operational_status is None:
        cleaned_data.pop("operational_status")

    if connection_status is None:
        cleaned_data.pop("connection_status")

    if internal_code:
        codes_seen.add(internal_code)

    if serial_number:
        serials_seen.add(serial_number)

    return ImportRowResult(
        row_number=row_number,
        data=cleaned_data,
        is_valid=not errors,
        errors=errors,
        warnings=warnings,
    )


# ==========================================================
# VISTA PREVIA
# ==========================================================

def preview_asset_import(uploaded_file):
    workbook, worksheet, headers = read_excel_file(
        uploaded_file
    )

    result = ImportResult()

    codes_seen = set()
    serials_seen = set()

    try:
        for row_number, row_values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if is_empty_row(row_values):
                continue

            raw_data = {
                headers[index]: value
                for index, value in enumerate(row_values)
                if index < len(headers) and headers[index]
            }

            row_result = validate_asset_row(
                row_number=row_number,
                raw_data=raw_data,
                codes_seen=codes_seen,
                serials_seen=serials_seen,
            )

            result.rows.append(row_result)
            result.total_rows += 1

            if row_result.is_valid:
                result.valid_rows += 1
            else:
                result.invalid_rows += 1

    finally:
        workbook.close()

    return result


# ==========================================================
# IMPORTACIÓN
# ==========================================================

def import_valid_assets(preview_result):
    imported_rows = 0
    skipped_rows = 0

    with transaction.atomic():
        for row_result in preview_result.rows:
            if not row_result.is_valid:
                skipped_rows += 1
                continue

            asset = Asset(
                **row_result.data
            )

            try:
                asset.full_clean()
                asset.save()
                imported_rows += 1

            except ValidationError as exc:
                row_result.is_valid = False

                if hasattr(exc, "message_dict"):
                    for field_errors in exc.message_dict.values():
                        row_result.errors.extend(field_errors)
                else:
                    row_result.errors.extend(exc.messages)

                skipped_rows += 1

    preview_result.imported_rows = imported_rows
    preview_result.skipped_rows = skipped_rows
    preview_result.invalid_rows = sum(
        1
        for row in preview_result.rows
        if not row.is_valid
    )
    preview_result.valid_rows = sum(
        1
        for row in preview_result.rows
        if row.is_valid
    )

    return preview_result


# ==========================================================
# GENERACIÓN DE PLANTILLA
# ==========================================================

def generate_asset_import_template():
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = "Activos"

    worksheet.append(TEMPLATE_COLUMNS)

    example_row = [
        "PC-ADM-001",
        "PAT-000001",
        "Computadora de escritorio",
        "Dell",
        "OptiPlex 7010",
        "SN123456789",
        "LOT-2026-001",
        "PC-ADM-001",
        "usuario@empresa.com",
        "Administración",
        "Oficina principal",
        "Operativo",
        "Sin información",
        "Windows 11 Pro",
        "192.168.1.25",
        "00:1A:2B:3C:4D:5E",
        "2026-01-15",
        "2029-01-15",
        "Proveedor Ejemplo",
        "Equipo nuevo",
    ]

    worksheet.append(example_row)

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(
        TEMPLATE_COLUMNS,
        start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            len(column_name) + 3,
            18,
        )

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    output.seek(0)

    return output
